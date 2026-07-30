"""
文件加载服务 —— 支持多格式文件的文本提取与分块

支持的格式：
  - 文本类：PDF / DOCX / TXT / Excel → 提取文字（Excel 转 Markdown 表格）+ 内嵌图片由视觉模型描述
  - 图片类：jpg/png/gif/bmp/webp     → qwen3-vl 视觉模型生成描述
  - 视频类：mp4/avi/mov/mkv          → 提取关键帧 → 视觉模型描述 → 合并
"""
import os
import time
import base64
import tempfile
import zipfile
import logging
import re
from pathlib import Path
from xml.etree import ElementTree as ET

import httpx
from langchain_community.document_loaders import PyPDFLoader, Docx2txtLoader, TextLoader
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_core.documents import Document
from langchain_ollama import OllamaEmbeddings

from rag_knowledge.config import Config
from rag_knowledge.ollama_http import OLLAMA_CLIENT_KWARGS
from rag_knowledge.models.document import FileCategory
from rag_knowledge.models.structured_document import (
    build_searchable_text,
    render_section_prefix,
)
from rag_knowledge.services.semantic_chunker import SemanticChunker, SemanticChunkingError
from rag_knowledge.services.unstructured_loader import UnstructuredChapterLoader, SUPPORTED_EXTS
from rag_knowledge.services.document_profiles import (
    DocumentProfile,
    apply_document_profile,
    finalize_profile_chunks,
    get_chunk_policy,
    normalize_document_profile,
)
from rag_knowledge.services.canonical_adapters import ADAPTER_EXTENSIONS, load_canonical_documents, load_canonical_result
from rag_knowledge.services.document_support import (
    IngestionDecision,
    FormatDisposition,
    classify_suffix,
    make_decision,
)
from dataclasses import dataclass
from rag_knowledge.services.section_chunk_merge import (
    CHUNKING_METHOD,
    TARGET_SOFT_MAX,
    apply_technical_manual_merge,
    reassign_chunk_adjacency,
)

logger = logging.getLogger(__name__)

# 各格式扩展名集合
TEXT_EXTS = {".docx", ".doc", ".txt", ".md", ".xls", ".xlsx", *ADAPTER_EXTENSIONS}
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".gif", ".bmp", ".webp"}
VIDEO_EXTS = {".mp4", ".avi", ".mov", ".mkv", ".flv", ".wmv"}
WORD_FIELD_RE = re.compile(r"(?im)^\s*(HYPERLINK|PAGEREF|TOC|MERGEFIELD|SEQ|REF)\b.*$")
TOC_LINE_RE = re.compile(r"^\s*(?:\d+(?:\.\d+)*)?\s*[\u4e00-\u9fffA-Za-z].*?(?:\.{2,}|…{2,}|\s{2,})\s*\d+\s*$")
URL_LINE_RE = re.compile(r"^\s*(?:https?://\S+|www\.\S+|\[[^\]]+\]\(https?://[^)]+\))\s*$", re.I)
VERSION_HINT_RE = re.compile(r"(?:v?\d+(?:\.\d+)+|rocky\s*\d+|centos\s*\d+|windows\s*server\s*\d+)", re.I)
TOC_MARKER_RE = re.compile(r"^(?:目录|contents?)(?:\s+\d+)?$", re.I)


@dataclass
class FileLoadResult:
    chunks: list[Document]
    category: str
    decisions: list[IngestionDecision]


class FileLoader:
    """多格式文件加载器"""

    def __init__(self):
        cfg = Config()
        self._document_profile_config = cfg
        self._chunk_size = cfg.chunk_size
        self._chunk_overlap = cfg.chunk_overlap
        self._semantic_chunking_enabled = cfg.semantic_chunking_enabled
        self._ollama_base = cfg.ollama_base_url
        self._vision_endpoint = cfg.vision_endpoint
        self._vision_model = cfg.vision_model
        self._extract_images = cfg.extract_embedded_images
        # 按自然语言段落边界切分，避免切断语义
        self._splitter = RecursiveCharacterTextSplitter(
            chunk_size=self._chunk_size,
            chunk_overlap=self._chunk_overlap,
            separators=["\n\n", "\n", "。", "！", "？", "；", "，", " ", ""],
        )
        self._semantic_chunker: SemanticChunker | None = None
        if self._semantic_chunking_enabled:
            try:
                embed_base = cfg.embedding_endpoint.resolved_base_url(cfg.ollama_base_url)
                if cfg.embedding_endpoint.normalized_provider() != "ollama":
                    logger.warning(
                        "embedding provider=%s 当前仅实现 OllamaEmbeddings；仍按 Ollama 协议访问 %s",
                        cfg.embedding_endpoint.provider,
                        embed_base,
                    )
                self._semantic_chunker = SemanticChunker(
                    embeddings=OllamaEmbeddings(
                        model=cfg.embedding_model,
                        base_url=embed_base,
                        client_kwargs=OLLAMA_CLIENT_KWARGS,
                    ),
                    fallback_splitter=self._splitter,
                    max_chunk_size=self._chunk_size,
                    min_chunk_size=cfg.semantic_min_chunk_size,
                    breakpoint_percentile=cfg.semantic_breakpoint_percentile,
                )
            except Exception as exc:
                logger.warning(
                    "semantic embedding client initialization failed; fixed splitter will be used: %s",
                    exc,
                )

        # unstructured 章节切片（可配置开关）
        self._use_unstructured = cfg.use_unstructured
        self._unstructured_loader: UnstructuredChapterLoader | None = None
        if self._use_unstructured:
            self._unstructured_loader = UnstructuredChapterLoader(
                chunk_size=self._chunk_size,
                chunk_overlap=0 if self._semantic_chunking_enabled else self._chunk_overlap,
                strategy=cfg.unstructured_strategy,
            )

    # ------------------------------------------------------------------
    # 对外接口
    # ------------------------------------------------------------------

    def load_with_decisions(
        self,
        file_path: str,
        *,
        document_profile: DocumentProfile | str | None = None,
    ) -> FileLoadResult:
        """
        加载并处理一个文件，产生分块与决策。
        """
        path = Path(file_path)
        suffix = path.suffix.lower()
        disp = classify_suffix(suffix)

        if disp.action == "excluded":
            dec = make_decision(path, status="excluded", reason_code=disp.reason_code)
            return FileLoadResult([], "text", [dec])

        if disp.action == "queued":
            dec = make_decision(path, status="queued", reason_code=disp.reason_code)
            category = FileCategory.TEXT
            if suffix in IMAGE_EXTS:
                category = FileCategory.IMAGE
            elif suffix in VIDEO_EXTS:
                category = FileCategory.VIDEO
            return FileLoadResult([], category, [dec])

        if suffix in TEXT_EXTS:
            chunks, decisions = self._load_text_with_decisions(file_path, document_profile=document_profile)
            return FileLoadResult(chunks, FileCategory.TEXT, decisions)
        elif suffix in IMAGE_EXTS:
            dec = make_decision(path, status="queued", reason_code="MEDIA_PROCESSING_DEFERRED")
            return FileLoadResult([], FileCategory.IMAGE, [dec])
        elif suffix in VIDEO_EXTS:
            dec = make_decision(path, status="queued", reason_code="MEDIA_PROCESSING_DEFERRED")
            return FileLoadResult([], FileCategory.VIDEO, [dec])
        else:
            dec = make_decision(path, status="excluded", reason_code="UNSUPPORTED_EXTENSION")
            return FileLoadResult([], FileCategory.TEXT, [dec])

    def load(
        self,
        file_path: str,
        *,
        document_profile: DocumentProfile | str | None = None,
    ) -> tuple[list[Document], str]:
        """
        兼容接口：内部调用 load_with_decisions，若存在文件级 queued/excluded 决策则抛出 ValueError。
        """
        res = self.load_with_decisions(file_path, document_profile=document_profile)
        if res.decisions:
            for d in res.decisions:
                if not d.locator:
                    raise ValueError(f"{d.reason_code}: {Path(file_path).name}")
        return res.chunks, res.category

    @staticmethod
    def detect_category(file_path: str) -> str | None:
        """根据后缀判断文件分类"""
        suffix = Path(file_path).suffix.lower()
        if suffix in TEXT_EXTS:
            return FileCategory.TEXT
        if suffix in IMAGE_EXTS:
            return FileCategory.IMAGE
        if suffix in VIDEO_EXTS:
            return FileCategory.VIDEO
        return None

    # ------------------------------------------------------------------
    # 文本类
    # ------------------------------------------------------------------

    def _load_text(
        self,
        file_path: str,
        *,
        document_profile: DocumentProfile | str | None = None,
    ) -> list[Document]:
        chunks, _ = self._load_text_with_decisions(file_path, document_profile=document_profile)
        return chunks

    def _load_text_with_decisions(
        self,
        file_path: str,
        *,
        document_profile: DocumentProfile | str | None = None,
    ) -> tuple[list[Document], list[IngestionDecision]]:
        path = Path(file_path)
        suffix = path.suffix.lower()
        profile = normalize_document_profile(document_profile)
        policy = get_chunk_policy(profile, config=getattr(self, "_document_profile_config", None))

        if suffix == ".doc":
            raise ValueError(f"LEGACY_DOC_REQUIRES_CONVERSION: {path.name}")

        if suffix in ADAPTER_EXTENSIONS:
            canonical_res = load_canonical_result(path)
            chunks = apply_document_profile(canonical_res.documents, profile, policy=policy)
            chunks = self._split_documents_preserving_blocks(chunks)
            chunks = self._post_process_chunks(chunks)
            return finalize_profile_chunks(chunks, profile, policy), canonical_res.decisions

        decisions = []
        if suffix == ".docx":
            try:
                with zipfile.ZipFile(file_path) as z:
                    media_files = [name for name in z.namelist() if name.startswith("word/media/")]
                    for idx, name in enumerate(media_files, 1):
                        decisions.append(make_decision(
                            path, status="queued", reason_code="EMBEDDED_MEDIA_PROCESSING_DEFERRED",
                            locator=f"docx:media:{idx}"
                        ))
            except Exception:
                pass

        if self._use_unstructured and suffix in SUPPORTED_EXTS:
            try:
                chunks = self._unstructured_loader.load(file_path)
                chunks = apply_document_profile(chunks, profile, policy=policy)
                chunks = self._split_documents_preserving_blocks(chunks)
            except Exception as e:
                logger.warning("unstructured 解析失败，回退旧解析方式 %s: %s", path.name, e)
                chunks = apply_document_profile(self._load_text_legacy(file_path), profile, policy=policy)
        else:
            chunks = apply_document_profile(self._load_text_legacy(file_path), profile, policy=policy)

        chunks = self._mark_table_context_chunks(chunks)
        chunks = self._post_process_chunks(chunks)
        return finalize_profile_chunks(chunks, profile, policy), decisions

    def _load_text_legacy(self, file_path: str) -> list[Document]:
        """旧版文本加载逻辑（固定长度切分）"""
        path = Path(file_path)
        suffix = path.suffix.lower()

        if suffix == ".pdf":
            loader = PyPDFLoader(str(path))
            docs = loader.load()
        elif suffix == ".docx":
            loader = Docx2txtLoader(str(path))
            docs = loader.load()
        elif suffix == ".doc":
            docs = self._load_old_doc(file_path)
        elif suffix in (".xls", ".xlsx"):
            docs = self._load_excel(file_path)
        else:
            loader = TextLoader(str(path), encoding="utf-8")
            docs = loader.load()

        for d in docs:
            d.metadata["source"] = path.name
            d.metadata["category"] = FileCategory.TEXT

        return self._split_documents_preserving_blocks(docs)

    def _load_old_doc(self, file_path: str) -> list[Document]:
        """提取旧版 Word .doc 文件（OLE2 二进制格式）中的文本

        支持 Unicode (UTF-16LE) 和 ANSI (GBK/系统代码页) 编码的文档。
        """
        path = Path(file_path)
        try:
            import olefile
        except ImportError:
            logger.warning("缺少 olefile 库，无法解析 .doc 文件: %s", path.name)
            return [Document(
                page_content=f"[需要安装 olefile 以支持旧版 .doc 文件] {path.name}",
                metadata={"source": path.name},
            )]

        try:
            ole = olefile.OleFileIO(file_path)
        except Exception as e:
            logger.error("无法打开 OLE2 容器 %s: %s", path.name, e)
            raise ValueError(f"无法解析旧版 .doc 文件: {path.name}")

        try:
            if not ole.exists("WordDocument"):
                raise ValueError(f"文件中未找到 WordDocument 流: {path.name}")

            stream = ole.openstream("WordDocument")
            data = stream.read()
            stream.close()

            import struct

            # 从 FIB 读取关键字段
            flags = struct.unpack_from("<H", data, 0x0A)[0] if len(data) > 12 else 0
            fc_min = struct.unpack_from("<I", data, 0x18)[0] if len(data) > 28 else 0
            ccp_text = struct.unpack_from("<I", data, 0x4C)[0] if len(data) > 80 else 0

            is_complex = bool(flags & 0x0200)  # fComplex 标志

            text = ""

            if not is_complex and ccp_text > 0:
                # ---- 非复杂格式：文本是连续块 ----
                start = max(fc_min, 0)
                try:
                    raw = data[start:start + ccp_text]
                    text = raw.decode("gbk").strip("\r\n\t\x00")
                except UnicodeDecodeError:
                    pass

                if not text:
                    try:
                        raw = data[start:start + ccp_text * 2]
                        text = raw.decode("utf-16-le", errors="ignore").strip("\r\n\t\x00")
                    except Exception:
                        pass

            if not text:
                candidates = []
                for start in sorted({fc_min, 0x200, 0x00}):
                    if start > 0 and start < len(data) - 100:
                        for enc, stride in [("gbk", 1), ("utf-16-le", 2), ("utf-8", 1)]:
                            raw_len = ccp_text * stride if ccp_text > 0 else min(len(data) - start, 50000)
                            if start + raw_len > len(data):
                                raw_len = len(data) - start
                            try:
                                decoded = data[start:start + raw_len].decode(enc, errors="replace")
                                good = sum(1 for c in decoded if c.isprintable() or c in "\n\r\t")
                                score = good / max(len(decoded), 1)
                                if score > 0.3:
                                    candidates.append((score, decoded))
                            except Exception:
                                continue

                if candidates:
                    candidates.sort(key=lambda x: -x[0])
                    text = candidates[0][1]

            if text:
                text = text.replace("�", "")
                text = "".join(c for c in text if c.isprintable() or c in "\n\r\t")
                text = "\n".join(line.strip() for line in text.split("\n") if line.strip())

            if not text:
                logger.warning("未能从 .doc 文件中提取到文本: %s", path.name)
                text = f"[未能提取文本的 Word 文档] {path.name}"

        except Exception as e:
            logger.error("解析 .doc 文件失败 %s: %s", path.name, e)
            raise
        finally:
            ole.close()

        doc = Document(
            page_content=text,
            metadata={"source": path.name},
        )
        return [doc]

    def _load_excel(self, file_path: str) -> list[Document]:
        """将 Excel 文件每个 sheet 转为 Markdown 表格格式的 Document。

        .xlsx 使用 openpyxl（data_only=True 读取计算后的值）；
        .xls  使用 xlrd（xlrd 2.x 仅支持旧版二进制格式）。
        """
        path = Path(file_path)
        suffix = path.suffix.lower()
        docs: list[Document] = []
        import hashlib

        snapshot_hash = hashlib.sha256(path.read_bytes()).hexdigest()

        def table_metadata(sheet_name: str, order: int) -> dict:
            element_id = f"el_excel_{snapshot_hash[:12]}_{order:04d}"
            return {
                "source": path.name,
                "sheet": sheet_name,
                "section_title": sheet_name,
                "section_path": sheet_name,
                "heading_level": 1,
                "element_order": order,
                "content_type": "table",
                "content_role": "table",
                "chunking_method": "table",
                "table_source": "excel",
                "element_id": element_id,
                "source_element_ids": [element_id],
                "source_raw_block_ids": [f"rb_excel_{snapshot_hash[:12]}_{order:04d}"],
                "source_document_id": snapshot_hash[:32],
                "source_snapshot_hash": snapshot_hash,
            }

        try:
            if suffix == ".xlsx":
                try:
                    import openpyxl  # noqa: PLC0415
                except ImportError:
                    logger.warning("缺少 openpyxl，无法解析 .xlsx: %s", path.name)
                    return [Document(
                        page_content=f"[需要安装 openpyxl 以支持 .xlsx 文件] {path.name}",
                        metadata={"source": path.name},
                    )]
                wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        continue
                    md = self._rows_to_markdown(rows)
                    if md.strip():
                        docs.append(Document(
                            page_content=md,
                            metadata=table_metadata(sheet_name, len(docs) + 1),
                        ))
                wb.close()

            elif suffix == ".xls":
                try:
                    import xlrd  # noqa: PLC0415
                except ImportError:
                    logger.warning("缺少 xlrd，无法解析 .xls: %s", path.name)
                    return [Document(
                        page_content=f"[需要安装 xlrd 以支持 .xls 文件] {path.name}",
                        metadata={"source": path.name},
                    )]
                wb = xlrd.open_workbook(str(path))
                for sheet in wb.sheets():
                    rows = [
                        tuple(sheet.cell_value(r, c) for c in range(sheet.ncols))
                        for r in range(sheet.nrows)
                    ]
                    if not rows:
                        continue
                    md = self._rows_to_markdown(rows)
                    if md.strip():
                        docs.append(Document(
                            page_content=md,
                            metadata=table_metadata(sheet.name, len(docs) + 1),
                        ))
            else:
                raise ValueError(f"不支持的 Excel 格式: {suffix}")

        except Exception as e:
            logger.error("解析 Excel 文件失败 %s: %s", path.name, e)
            raise

        if not docs:
            logger.warning("Excel 文件内容为空: %s", path.name)
            docs.append(Document(
                page_content=f"[空 Excel 文件] {path.name}",
                metadata={"source": path.name},
            ))
        return docs

    @staticmethod
    def _rows_to_markdown(rows: list) -> str:
        """将二维行列数据转换为 Markdown 表格字符串。

        第一行作为表头，管道符 `|` 自动转义，None 值转为空字符串。
        列数不一致时补空列以对齐。
        """
        if not rows:
            return ""

        def _cell(v) -> str:
            if v is None:
                return ""
            return str(v).replace("|", "\\|").strip()

        header = rows[0]
        cols = len(header)
        if cols == 0:
            return ""

        lines = []
        lines.append("| " + " | ".join(_cell(c) for c in header) + " |")
        lines.append("| " + " | ".join("---" for _ in range(cols)) + " |")
        for row in rows[1:]:
            padded = list(row) + [None] * max(0, cols - len(row))
            lines.append("| " + " | ".join(_cell(c) for c in padded[:cols]) + " |")
        return "\n".join(lines)

    def _post_process_chunks(self, chunks: list[Document]) -> list[Document]:
        """统一清洗分块内容，过滤低信息噪声块。"""
        cleaned_chunks: list[Document] = []
        for chunk in chunks:
            content_type = chunk.metadata.get("content_type")
            if content_type in ("code", "table", "heading"):
                # 结构保护块跳过低信息过滤和空格折叠，但做首尾空白清理
                cleaned = chunk.page_content.strip()
                if not cleaned:
                    continue
                chunk.page_content = cleaned
                cleaned_chunks.append(chunk)
            else:
                cleaned = self._sanitize_text(chunk.page_content)
                if (
                    not cleaned
                    or self._is_toc_marker_chunk(cleaned)
                    or self._is_low_information(cleaned)
                ):
                    continue
                chunk.page_content = cleaned
                cleaned_chunks.append(chunk)
        return cleaned_chunks

    @staticmethod
    def _sanitize_text(text: str) -> str:
        """清洗 Word 域代码、无效链接和多余空白。"""
        if not text:
            return ""

        cleaned = text.replace("\x13", " ").replace("\x14", " ").replace("\x15", " ")
        cleaned = WORD_FIELD_RE.sub("", cleaned)
        cleaned = re.sub(r"(?im)^\s*TOC\s+\\[^\n]*$", "", cleaned)
        cleaned = re.sub(r"(?im)^\s*PAGEREF\s+[^\n]*$", "", cleaned)
        cleaned = re.sub(r"(?im)^\s*HYPERLINK\s+\"[^\n]*$", "", cleaned)
        cleaned = re.sub(r"[ \t]+", " ", cleaned)
        cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
        return cleaned.strip()

    @staticmethod
    def _is_toc_marker_chunk(text: str) -> bool:
        body = text.strip()
        if body.startswith("# ") and "\n\n" in body:
            body = body.split("\n\n", 1)[1].strip()
        return bool(TOC_MARKER_RE.fullmatch(body))

    @staticmethod
    def _is_low_information(text: str) -> bool:
        """过滤目录块、纯链接块和极短噪声块。"""
        stripped = text.strip()
        if not stripped:
            return True

        compact = re.sub(r"[\W_]+", "", stripped, flags=re.UNICODE)
        if len(compact) < 8:
            return True

        lines = [line.strip() for line in stripped.splitlines() if line.strip()]
        if not lines:
            return True

        toc_hits = sum(1 for line in lines if TOC_LINE_RE.match(line))
        if len(lines) >= 2 and toc_hits / len(lines) >= 0.6:
            return True

        url_hits = sum(1 for line in lines if URL_LINE_RE.match(line))
        if url_hits == len(lines):
            return True

        if len(lines) == 1 and len(compact) < 16 and not VERSION_HINT_RE.search(stripped):
            return True

        non_space_chars = [char for char in stripped if not char.isspace()]
        if len(non_space_chars) >= 24:
            unexpected_script_count = sum(
                1 for char in non_space_chars if FileLoader._is_unexpected_script_char(char)
            )
            if unexpected_script_count:
                return True
            readable = sum(1 for char in non_space_chars if FileLoader._is_readable_text_char(char))
            if readable / len(non_space_chars) < 0.35:
                return True
            longest_unreadable_run = 0
            current_unreadable_run = 0
            for char in stripped:
                if char.isspace() or FileLoader._is_readable_text_char(char):
                    current_unreadable_run = 0
                    continue
                current_unreadable_run += 1
                longest_unreadable_run = max(longest_unreadable_run, current_unreadable_run)
            if longest_unreadable_run >= 12:
                return True

        return False

    @staticmethod
    def _is_readable_text_char(char: str) -> bool:
        if not char:
            return False
        if char.isascii():
            return char.isalnum() or char in ".,;:!?()[]{}<>/-_#%&*+=@'\"\\|`~$"
        if "\u3400" <= char <= "\u9fff":
            return True
        if char in "，。；：！？、（）《》【】“”‘’—…￥":
            return True
        return False

    @staticmethod
    def _is_unexpected_script_char(char: str) -> bool:
        if not char:
            return False
        codepoint = ord(char)
        return (
            0x0590 <= codepoint <= 0x08FF  # Hebrew, Arabic and related scripts
            or 0x0900 <= codepoint <= 0x0D7F  # Indic scripts, including Malayalam
            or 0x0E00 <= codepoint <= 0x0FFF  # Thai, Lao and Tibetan
            or 0x1000 <= codepoint <= 0x10FF  # Myanmar and Georgian
        )

    # ------------------------------------------------------------------
    # 内嵌图片提取（Word / PDF 中的图片 → 视觉模型描述）
    # ------------------------------------------------------------------

    def _describe_embedded_images(self, file_path: str, max_images: int = 5) -> list[str]:
        """提取文档内嵌图片并调用视觉模型描述，返回描述文本列表"""
        suffix = Path(file_path).suffix.lower()
        if suffix == ".docx":
            return self._describe_docx_images(file_path, max_images)
        elif suffix == ".pdf":
            return self._describe_pdf_images(file_path, max_images)
        return []

    def _describe_docx_images(self, file_path: str, max_images: int) -> list[str]:
        """从 DOCX（zip 包）中提取内嵌图片 → 视觉模型描述"""
        path = Path(file_path)
        descs: list[str] = []
        try:
            with zipfile.ZipFile(file_path) as z:
                rels_path = "word/_rels/document.xml.rels"
                if rels_path not in z.namelist():
                    return []

                # 解析关系文件，映射 rId → 图片路径
                rels_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
                rels_xml = ET.parse(z.open(rels_path))
                rid_map: dict[str, str] = {}
                for rel in rels_xml.findall(f"{{{rels_ns}}}Relationship"):
                    rid = rel.get("Id", "")
                    target = rel.get("Target", "")
                    img_exts = (".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp")
                    if any(target.lower().endswith(ext) for ext in img_exts):
                        full_path = f"word/{target}"
                        if full_path in z.namelist():
                            rid_map[rid] = full_path

                if not rid_map:
                    return []

                rids = list(rid_map.keys())[:max_images]
                for i, rid in enumerate(rids):
                    img_path = rid_map[rid]
                    img_data = z.open(img_path).read()
                    ext = Path(img_path).suffix
                    tmp = tempfile.NamedTemporaryFile(suffix=ext, delete=False)
                    tmp.write(img_data)
                    tmp.close()
                    try:
                        desc = self._call_vision(tmp.name)
                        if desc:
                            descs.append(f"[{path.name} 图片{i + 1}]\n{desc.strip()}")
                    finally:
                        if os.path.exists(tmp.name):
                            os.unlink(tmp.name)
        except Exception as e:
            logger.warning("提取 DOCX 内嵌图片失败 %s: %s", path.name, e)
        return descs

    def _describe_pdf_images(self, file_path: str, max_images: int) -> list[str]:
        """从 PDF 中提取内嵌图片 → 视觉模型描述（需要 PyMuPDF）"""
        path = Path(file_path)
        descs: list[str] = []
        try:
            import fitz  # PyMuPDF
        except ImportError:
            logger.warning("缺少 PyMuPDF，无法提取 PDF 内嵌图片: %s", path.name)
            return []

        try:
            doc = fitz.open(file_path)
            count = 0
            for page_num in range(len(doc)):
                if count >= max_images:
                    break
                images = doc[page_num].get_images(full=True)
                for img_info in images:
                    if count >= max_images:
                        break
                    xref = img_info[0]
                    base_image = doc.extract_image(xref)
                    img_bytes = base_image["image"]
                    ext = base_image["ext"]

                    tmp = tempfile.NamedTemporaryFile(suffix=f".{ext}", delete=False)
                    tmp.write(img_bytes)
                    tmp.close()
                    try:
                        desc = self._call_vision(tmp.name)
                        if desc:
                            count += 1
                            descs.append(f"[{path.name} 图片{count}]\n{desc.strip()}")
                    finally:
                        if os.path.exists(tmp.name):
                            os.unlink(tmp.name)
            doc.close()
        except Exception as e:
            logger.warning("提取 PDF 内嵌图片失败 %s: %s", path.name, e)
        return descs

    # ------------------------------------------------------------------
    # 图片类
    # ------------------------------------------------------------------

    def _load_image(self, file_path: str) -> list[Document]:
        """图片 → 压缩 → 视觉模型描述 → 分块"""
        path = Path(file_path)
        t0 = time.time()
        logger.info("处理图片: %s", path.name)

        try:
            processed = self._resize_image(file_path, max_px=1024)
        except Exception as e:
            logger.error("图片压缩失败 %s: %s", path.name, e)
            processed = file_path

        try:
            desc = self._call_vision(processed)
        finally:
            if processed != file_path and os.path.exists(processed):
                os.unlink(processed)

        content = desc.strip() or f"[未能生成描述的图片] {path.name}"
        doc = Document(
            page_content=content,
            metadata={"source": path.name, "category": FileCategory.IMAGE, "file_path": str(path)},
        )
        result = self._splitter.split_documents([doc])
        elapsed = time.time() - t0
        logger.info("图片处理完成: %s | %d 块 | %.2fs", path.name, len(result), elapsed)
        return result

    # ------------------------------------------------------------------
    # 视频类
    # ------------------------------------------------------------------

    def _load_video(self, file_path: str) -> list[Document]:
        """视频 → 提取 N 个关键帧 → 逐一描述 → 合并分块"""
        path = Path(file_path)
        t0 = time.time()
        logger.info("处理视频: %s", path.name)

        descs = self._extract_frames(file_path)
        if not descs:
            content = f"[视频文件] {path.name}（无法提取帧）"
        else:
            parts = "\n\n".join(
                f"--- 第 {i + 1} 帧 ---\n{t}" for i, t in enumerate(descs)
            )
            content = f"视频 [{path.name}] 内容描述：\n{parts}"

        doc = Document(
            page_content=content,
            metadata={"source": path.name, "category": FileCategory.VIDEO, "file_path": str(path)},
        )
        result = self._splitter.split_documents([doc])
        elapsed = time.time() - t0
        logger.info("视频处理完成: %s | %d 块 | %.2fs", path.name, len(result), elapsed)
        return result

    def _extract_frames(self, video_path: str, n: int = 3) -> list[str]:
        """用 OpenCV 提取均匀分布的 n 帧并逐帧描述"""
        try:
            import cv2
        except ImportError:
            logger.warning("缺少 opencv-python-headless，跳过视频: %s", video_path)
            return ["[需要安装 opencv-python-headless 以支持视频处理]"]

        cap = cv2.VideoCapture(video_path)
        if not cap.isOpened():
            return []

        total = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
        fps = cap.get(cv2.CAP_PROP_FPS)
        if total <= 0:
            cap.release()
            return []

        descs = []
        for i in range(1, n + 1):
            target = int(total * i / (n + 1))
            cap.set(cv2.CAP_PROP_POS_FRAMES, target)
            ok, frame = cap.read()
            if not ok:
                continue

            tmp = tempfile.NamedTemporaryFile(suffix=".jpg", delete=False)
            tmp_path = tmp.name
            tmp.close()
            try:
                cv2.imwrite(tmp_path, frame)
                desc = self._call_vision(tmp_path)
                if desc and desc.strip():
                    ts = target / fps if fps > 0 else 0
                    descs.append(f"[{int(ts // 60):02d}:{int(ts % 60):02d}] {desc.strip()}")
            finally:
                if os.path.exists(tmp_path):
                    os.unlink(tmp_path)

        cap.release()
        if fps > 0:
            descs.insert(0, f"时长 {int(total / fps // 60)} 分 {int(total / fps % 60)} 秒")
        return descs

    # ------------------------------------------------------------------
    # 视觉模型调用（qwen3-vl）
    # ------------------------------------------------------------------

    def _call_vision(self, image_path: str) -> str:
        """调用视觉模型描述图片（ollama / openai / google），失败时返回空字符串"""
        try:
            with open(image_path, "rb") as f:
                b64 = base64.b64encode(f.read()).decode("utf-8")
        except Exception as e:
            logger.warning("读取图片失败 %s: %s", image_path, e)
            return ""

        from rag_knowledge.llm_http import ModelEndpoint, chat_vision

        ep = self._vision_endpoint
        if self._vision_model and self._vision_model != ep.model:
            ep = ModelEndpoint(
                role=ep.role,
                provider=ep.provider,
                model=self._vision_model,
                base_url=ep.base_url,
                api_key_env=ep.api_key_env,
            )
        suffix = Path(image_path).suffix.lower()
        mime = {
            ".png": "image/png",
            ".gif": "image/gif",
            ".webp": "image/webp",
            ".bmp": "image/bmp",
        }.get(suffix, "image/jpeg")
        try:
            return chat_vision(
                ep,
                "请详细描述这张图片的内容，包括文字、物体、人物、场景、颜色等所有可见信息",
                b64,
                default_ollama=self._ollama_base,
                mime_type=mime,
                timeout=180.0,
            )
        except httpx.TimeoutException:
            logger.warning("视觉模型超时: %s（图片可能过大）", Path(image_path).name)
            return ""
        except httpx.HTTPStatusError as e:
            logger.warning("视觉模型返回错误 %s: %s", e.response.status_code, Path(image_path).name)
            return ""
        except Exception as e:
            logger.warning("视觉模型调用失败 %s: %s", Path(image_path).name, e)
            return ""

    # ------------------------------------------------------------------
    # 工具方法
    # ------------------------------------------------------------------

    @staticmethod
    def _resize_image(image_path: str, max_px: int = 1024) -> str:
        try:
            from PIL import Image
        except ImportError:
            return image_path

        img = Image.open(image_path)

        if img.width <= max_px and img.height <= max_px:
            return image_path

        img.thumbnail((max_px, max_px), Image.LANCZOS)

        if img.mode in ("RGBA", "LA", "P"):
            img = img.convert("RGB")

        tmp = tempfile.NamedTemporaryFile(suffix=".jpg", delete=False)
        img.save(tmp.name, "JPEG", quality=85)
        return tmp.name

    def _split_documents_preserving_blocks(self, docs: list[Document]) -> list[Document]:
        """对文档列表进行结构保护切分"""
        out_docs = []
        for doc in docs:
            out_docs.extend(self._split_markdown_preserving_blocks(doc))
        return [self._finalize_structured_chunk(doc) for doc in out_docs]

    @staticmethod
    def _mark_table_context_chunks(chunks: list[Document]) -> list[Document]:
        for index, chunk in enumerate(chunks):
            meta = chunk.metadata or {}
            if meta.get("content_type") != "text":
                continue
            path = str(meta.get("section_path") or "")
            related_tables = []
            has_adjacent_table = False
            for neighbor_index in (index - 1, index + 1):
                if not 0 <= neighbor_index < len(chunks):
                    continue
                neighbor_meta = chunks[neighbor_index].metadata or {}
                if (
                    neighbor_meta.get("content_type") == "table"
                    and str(neighbor_meta.get("section_path") or "") == path
                ):
                    has_adjacent_table = True
                    table_id = str(neighbor_meta.get("table_id") or "")
                    if table_id:
                        related_tables.append(table_id)
            if has_adjacent_table:
                updated = dict(meta)
                updated["table_context"] = True
                updated["related_table_ids"] = list(dict.fromkeys(related_tables))
                chunk.metadata = updated
        return chunks

    def _split_markdown_preserving_blocks(self, doc: Document) -> list[Document]:
        """将单个文档（Markdown/Excel 表格/普通文本）按结构保护切分"""
        metadata = doc.metadata.copy()
        content = doc.page_content

        # Round 0C: keep bounded merge windows, but split oversized source elements.
        if (
            metadata.get("chunking_method") == CHUNKING_METHOD
            and metadata.get("content_type") not in {"table", "code", "embedded_image"}
        ):
            section_path = str(metadata.get("section_path") or "")
            section_parts = [part.strip() for part in section_path.split(">") if part.strip()]
            prefix = render_section_prefix(section_parts)
            rendered_len = len(content) + (len(prefix) + 2 if prefix else 0)
            if rendered_len <= TARGET_SOFT_MAX:
                return [doc]
            return self._split_documents_with_method([doc], CHUNKING_METHOD)

        # 如果已经是 Excel 转换过来的表格，或者元数据中已经指定 content_type == "table"，
        # 且整个内容就是个 Markdown 表格，直接按表格规则处理。
        if metadata.get("content_type") == "table" or metadata.get("table_source") == "excel":
            return self._split_markdown_table(content, metadata)

        # 否则，扫描并提取其中的 block (tables, fenced code blocks, and plain text)
        blocks = self._extract_markdown_blocks(content)

        out_docs = []
        for block in blocks:
            b_type = block["type"]
            b_text = block["text"]
            b_metadata = metadata.copy()

            if b_type == "table":
                out_docs.extend(self._split_markdown_table(b_text, b_metadata))
            elif b_type == "code":
                out_docs.extend(self._split_fenced_code_block(b_text, b_metadata))
            else:
                if not b_text.strip():
                    continue
                temp_doc = Document(page_content=b_text, metadata=b_metadata)
                out_docs.extend(self._split_plain_text_block(temp_doc))

        return out_docs

    @staticmethod
    def _finalize_structured_chunk(doc: Document) -> Document:
        """Render section prefixes and stable searchable_text for final stored chunks."""
        metadata = dict(doc.metadata or {})
        body = (doc.page_content or "").strip()
        if not body:
            doc.metadata = metadata
            return doc

        section_path = metadata.get("section_path")
        if isinstance(section_path, str):
            section_parts = [part.strip() for part in section_path.split(">") if part.strip()]
        elif isinstance(section_path, (list, tuple)):
            section_parts = [str(part).strip() for part in section_path if str(part).strip()]
            metadata["section_path"] = " > ".join(section_parts)
        else:
            section_parts = []
            metadata.setdefault("section_path", "")

        prefix = render_section_prefix(section_parts)
        doc.page_content = f"{prefix}\n\n{body}" if prefix else body
        metadata["searchable_text"] = build_searchable_text(
            section_parts,
            body,
            metadata.get("content_type"),
        )
        doc.metadata = metadata
        return doc

    def _split_plain_text_block(self, doc: Document) -> list[Document]:
        if getattr(self, "_semantic_chunking_enabled", False) and getattr(self, "_semantic_chunker", None) is not None:
            try:
                return self._semantic_chunker.split_document(doc)
            except SemanticChunkingError as exc:
                logger.warning(
                    "semantic chunking unavailable for %s, falling back to fixed splitter: %s",
                    doc.metadata.get("source", ""),
                    exc,
                )
            except Exception as exc:
                logger.warning(
                    "semantic chunking failed for %s, falling back to fixed splitter: %s",
                    doc.metadata.get("source", ""),
                    exc,
                )
        return self._split_documents_with_method([doc], "fixed_fallback")

    def _split_documents_with_method(self, docs: list[Document], method: str) -> list[Document]:
        chunks = self._splitter.split_documents(docs)
        for chunk in chunks:
            chunk.metadata["chunking_method"] = method
        return chunks

    def _extract_markdown_blocks(self, text: str) -> list[dict]:
        """
        扫描 Markdown 文本，识别出普通文本、Markdown 表格块和 fenced code block。
        返回列表：[{"type": "text"|"table"|"code", "text": str}]
        """
        lines = text.splitlines()
        blocks = []
        i = 0
        n = len(lines)
        current_text_lines = []

        def flush_text():
            if current_text_lines:
                blocks.append({
                    "type": "text",
                    "text": "\n".join(current_text_lines)
                })
                current_text_lines.clear()

        while i < n:
            line = lines[i]
            stripped = line.strip()

            # 1. 检查是否是 fenced code block 的开始
            if stripped.startswith("```"):
                flush_text()
                code_block_lines = [line]
                i += 1
                while i < n:
                    inner_line = lines[i]
                    code_block_lines.append(inner_line)
                    if inner_line.strip().startswith("```"):
                        i += 1
                        break
                    i += 1
                blocks.append({
                    "type": "code",
                    "text": "\n".join(code_block_lines)
                })
                continue

            # 2. 检查是否是 Markdown 表格的开始
            if i + 1 < n and "|" in line and "|" in lines[i + 1] and re.match(r"^\s*\|?\s*:?-{3,}:?\s*(\|\s*:?-{3,}:?\s*)*\|?\s*$", lines[i + 1]):
                flush_text()
                table_lines = []
                while i < n:
                    tbl_line = lines[i]
                    tbl_stripped = tbl_line.strip()
                    if tbl_stripped.startswith("```"):
                        break
                    if "|" not in tbl_line:
                        break
                    table_lines.append(tbl_line)
                    i += 1
                blocks.append({
                    "type": "table",
                    "text": "\n".join(table_lines)
                })
                continue

            # 3. 普通文本行
            current_text_lines.append(line)
            i += 1

        flush_text()
        return blocks

    def _split_markdown_table(self, table_text: str, metadata: dict) -> list[Document]:
        """
        对 Markdown 表格进行分块。
        如果表格长度 <= chunk_size，作为一个 chunk。
        否则，按完整行切分，每个 chunk 都包含表头（前两行）并更新 row_start/row_end 元数据。
        """
        import hashlib
        table_id = f"table_{hashlib.md5(table_text.encode('utf-8')).hexdigest()[:8]}"

        lines = [line for line in table_text.splitlines() if line.strip()]
        if len(lines) < 2:
            return [Document(
                page_content=table_text,
                metadata={
                    **metadata,
                    "content_type": "table",
                    "chunking_method": "table",
                    "table_id": table_id,
                    "row_start": 1,
                    "row_end": len(lines),
                }
            )]

        header_row = lines[0]
        separator_row = lines[1]
        data_rows = lines[2:]

        if not data_rows:
            return [Document(
                page_content=table_text,
                metadata={
                    **metadata,
                    "content_type": "table",
                    "chunking_method": "table",
                    "table_id": table_id,
                    "row_start": 0,
                    "row_end": 0,
                }
            )]

        keep_whole_docx_table = (
            str(metadata.get("source", "")).lower().endswith(".docx")
            and len(table_text) <= max(self._chunk_size * 4, 4096)
        )

        if len(table_text) <= self._chunk_size or keep_whole_docx_table:
            return [Document(
                page_content=table_text,
                metadata={
                    **metadata,
                    "content_type": "table",
                    "chunking_method": "table",
                    "table_id": table_id,
                    "row_start": 1,
                    "row_end": len(data_rows),
                }
            )]

        chunks = []
        header_prefix = f"{header_row}\n{separator_row}\n"
        prefix_len = len(header_prefix)

        current_chunk_rows = []
        current_len = prefix_len
        start_row_idx = 1

        for idx, row in enumerate(data_rows, start=1):
            row_len = len(row) + 1
            if current_len + row_len > self._chunk_size and current_chunk_rows:
                chunk_text = header_prefix + "\n".join(current_chunk_rows)
                chunks.append(Document(
                    page_content=chunk_text,
                    metadata={
                        **metadata,
                        "content_type": "table",
                        "chunking_method": "table",
                        "table_id": table_id,
                        "row_start": start_row_idx,
                        "row_end": idx - 1,
                    }
                ))
                current_chunk_rows = [row]
                current_len = prefix_len + len(row)
                start_row_idx = idx
            else:
                current_chunk_rows.append(row)
                current_len += row_len

        if current_chunk_rows:
            chunk_text = header_prefix + "\n".join(current_chunk_rows)
            chunks.append(Document(
                page_content=chunk_text,
                metadata={
                    **metadata,
                    "content_type": "table",
                    "chunking_method": "table",
                    "table_id": table_id,
                    "row_start": start_row_idx,
                    "row_end": len(data_rows),
                }
            ))

        return chunks

    def _split_fenced_code_block(self, code_text: str, metadata: dict) -> list[Document]:
        """
        对 fenced code block 进行分块。
        如果长度 <= chunk_size，作为一个 chunk。
        否则，按完整行切分，每个 chunk 都用 ```lang 开头和 ``` 结尾包围，
        并添加 code_block_index, part_index, language, content_type 等元数据。
        """
        import hashlib
        code_hash = hashlib.md5(code_text.encode('utf-8')).hexdigest()[:8]
        code_block_id = f"code_{code_hash}"

        lines = code_text.splitlines()
        if len(lines) < 2 or not lines[0].strip().startswith("```") or not lines[-1].strip().startswith("```"):
            return [Document(
                page_content=code_text,
                metadata={
                    **metadata,
                    "content_type": "code",
                    "chunking_method": "code",
                    "language": "",
                    "code_block_index": code_block_id,
                    "part_index": 1,
                }
            )]

        first_line_stripped = lines[0].strip()
        lang_match = re.match(r"^```\s*(\w+)?", first_line_stripped)
        language = lang_match.group(1) if lang_match and lang_match.group(1) else ""

        code_lines = lines[1:-1]

        if len(code_text) <= self._chunk_size:
            return [Document(
                page_content=code_text,
                metadata={
                    **metadata,
                    "content_type": "code",
                    "chunking_method": "code",
                    "language": language,
                    "code_block_index": code_block_id,
                    "part_index": 1,
                }
            )]

        chunks = []
        prefix = f"```{language}\n"
        suffix = "\n```"
        prefix_len = len(prefix)
        suffix_len = len(suffix)

        current_chunk_lines = []
        current_len = prefix_len + suffix_len
        part_idx = 1

        for idx, line in enumerate(code_lines, start=1):
            line_len = len(line) + 1
            if current_len + line_len > self._chunk_size and current_chunk_lines:
                chunk_text = prefix + "\n".join(current_chunk_lines) + suffix
                chunks.append(Document(
                    page_content=chunk_text,
                    metadata={
                        **metadata,
                        "content_type": "code",
                        "chunking_method": "code",
                        "language": language,
                        "code_block_index": code_block_id,
                        "part_index": part_idx,
                    }
                ))
                part_idx += 1
                current_chunk_lines = [line]
                current_len = prefix_len + suffix_len + len(line)
            else:
                current_chunk_lines.append(line)
                current_len += line_len

        if current_chunk_lines:
            chunk_text = prefix + "\n".join(current_chunk_lines) + suffix
            chunks.append(Document(
                page_content=chunk_text,
                metadata={
                    **metadata,
                    "content_type": "code",
                    "chunking_method": "code",
                    "language": language,
                    "code_block_index": code_block_id,
                    "part_index": part_idx,
                }
            ))

        return chunks

    def close(self):
        return
